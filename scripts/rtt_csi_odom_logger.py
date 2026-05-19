#!/usr/bin/env python3
"""
rtt_csi_odom_logger.py — PHOENIX SAR Robot RTT/CSI Synchronized Data Logger

Reads FTM + CSI serial output from the onboard XIAO ESP32-S3 (ranging initiator)
and synchronizes each measurement burst with:
  - ROS clock timestamp (captured at BURST_START)
  - /diff_cont/odom odometry (latest available when data arrives)
  - map -> base_link TF pose (latest available from SLAM/Nav2)

Output (saved to ~/phoenix_data/run_YYYYMMDD_HHMMSS/):
  ftm_frames.csv     — one row per FTM frame
  csi_packets.csv    — one row per CSI packet
  burst_summary.csv  — one row per burst (aggregated)
  metadata.json      — run metadata
  ../run_YYYYMMDD_HHMMSS.xlsx  — Excel workbook with all sheets

Serial protocol from firmware (115200 baud):
  BURST_START,<seq>,<n_frames>,<label>
  FTM_F,<seq>,<frame_idx>,<rtt_ps>,<t1_ps>,<t2_ps>,<t3_ps>,<t4_ps>,<rssi_dbm>,<label>
  CSI,<seq>,<rssi_dbm>,<noise_floor_dbm>,<n_sub>,<amp0>,...,<ampN-1>,<label>
  LABEL,<label>

Usage — as a ROS2 node:
  ros2 run articubot_one rtt_csi_odom_logger \
    --ros-args \
    -p serial_port:=/dev/ttyACM0 \
    -p baud_rate:=115200 \
    -p output_dir:=/home/phoenix/phoenix_data \
    -p odom_topic:=/diff_cont/odom \
    -p fixed_frame:=map \
    -p base_frame:=base_link \
    -p notes:="LOS baseline test, AP at door"

Usage — as a plain Python script (no colcon build needed):
  python3 scripts/rtt_csi_odom_logger.py \
    --serial-port /dev/ttyACM0 \
    --baud 115200 \
    --output-dir ~/phoenix_data

To send a label to the XIAO (changes label for subsequent bursts):
  # Via picocom (type label then Enter):
  picocom -b 115200 /dev/ttyACM0

  # Via python one-liner (non-blocking, can run while logger is active
  # if logger is opened in write mode too — but usually just restart logger):
  python3 -c "import serial,time; s=serial.Serial('/dev/ttyACM0',115200); s.write(b'LOS_DYNAMIC\\r\\n'); time.sleep(0.1); s.close()"

Dependencies:
  pip install pyserial openpyxl
  # or:
  sudo apt install python3-serial python3-openpyxl
"""

import math
import os
import csv
import json
import sys
import threading
import argparse
from collections import defaultdict
from datetime import datetime, timezone

# Check for pyserial
try:
    import serial
except ImportError:
    print("ERROR: pyserial not installed. Run: pip install pyserial")
    sys.exit(1)

# Check for openpyxl (optional but preferred)
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not installed. CSVs will be saved but no Excel file.")
    print("         To enable Excel: pip install openpyxl")

# ROS 2 imports (optional — script works without ROS if run standalone)
try:
    import rclpy
    from rclpy.node import Node
    from rclpy.duration import Duration
    from nav_msgs.msg import Odometry
    import tf2_ros
    ROS_AVAILABLE = True
except ImportError:
    ROS_AVAILABLE = False
    print("WARNING: rclpy not found. Running without ROS (no odom/TF sync).")

SPEED_OF_LIGHT = 299_792_458.0  # m/s


# =============================================================================
# Data logger core — handles serial parsing and file writing
# =============================================================================

class DataLogger:
    """Handles serial parsing, data storage, and file writing."""

    FTM_HEADER = [
        'seq', 'frame_idx', 'label',
        'ros_burst_time_sec', 'ros_burst_time_nanosec',
        'rtt_ps', 'distance_m',
        't1_ps', 't2_ps', 't3_ps', 't4_ps',
        'rssi_dbm',
        'odom_stamp_sec', 'odom_stamp_nanosec',
        'odom_x', 'odom_y', 'odom_yaw',
        'map_x', 'map_y', 'map_yaw',
    ]

    CSI_HEADER = (
        ['seq', 'label',
         'ros_burst_time_sec', 'ros_burst_time_nanosec',
         'rssi_dbm', 'noise_floor_dbm', 'n_sub']
        + [f'amp_{i}' for i in range(52)]
        + ['odom_stamp_sec', 'odom_stamp_nanosec',
           'odom_x', 'odom_y', 'odom_yaw',
           'map_x', 'map_y', 'map_yaw']
    )

    BURST_HEADER = [
        'seq', 'label',
        'ros_burst_time_sec', 'ros_burst_time_nanosec',
        'n_frames', 'n_ftm_rows', 'n_csi_rows',
        'mean_rtt_ps', 'mean_distance_m', 'mean_rssi_dbm',
        'odom_x', 'odom_y', 'odom_yaw',
        'map_x', 'map_y', 'map_yaw',
    ]

    def __init__(self, run_dir, get_ros_time_fn, get_odom_fn, get_map_pose_fn):
        self.run_dir = run_dir
        self.get_ros_time = get_ros_time_fn   # () -> (sec, nanosec) or (0, 0)
        self.get_odom = get_odom_fn           # () -> (sec, ns, x, y, yaw) or Nones
        self.get_map_pose = get_map_pose_fn   # () -> (x, y, yaw) or Nones

        self._lock = threading.Lock()

        self.seq_burst_time = {}    # seq -> (sec, nanosec)
        self.seq_label = {}         # seq -> str
        self.seq_n_frames = {}      # seq -> int
        self.current_label = 'UNKNOWN'

        # Accumulator for burst summary
        self._burst_acc = defaultdict(lambda: {
            'ftm_count': 0, 'csi_count': 0,
            'rtt_list': [], 'rssi_list': [],
            'odom': (None,) * 5, 'map': (None,) * 3,
            'n_frames': 0,
        })

        os.makedirs(run_dir, exist_ok=True)
        self._open_csv_files()

    def _open_csv_files(self):
        self._ftm_f = open(os.path.join(self.run_dir, 'ftm_frames.csv'), 'w', newline='')
        self._csi_f = open(os.path.join(self.run_dir, 'csi_packets.csv'), 'w', newline='')
        self._burst_f = open(os.path.join(self.run_dir, 'burst_summary.csv'), 'w', newline='')

        self._ftm_w = csv.writer(self._ftm_f)
        self._csi_w = csv.writer(self._csi_f)
        self._burst_w = csv.writer(self._burst_f)

        self._ftm_w.writerow(self.FTM_HEADER)
        self._csi_w.writerow(self.CSI_HEADER)
        self._burst_w.writerow(self.BURST_HEADER)

    def parse_line(self, line: str):
        parts = [p.strip() for p in line.split(',')]
        if not parts or not parts[0]:
            return

        tag = parts[0]

        if tag == 'BURST_START':
            self._handle_burst_start(parts)
        elif tag == 'FTM_F':
            self._handle_ftm(parts)
        elif tag == 'CSI':
            self._handle_csi(parts)
        elif tag == 'LABEL':
            self._handle_label(parts)

    def _handle_burst_start(self, parts):
        # BURST_START,<seq>,<n_frames>,<label>
        if len(parts) < 4:
            return
        seq = int(parts[1])
        n_frames = int(parts[2])
        label = parts[3]

        ros_time = self.get_ros_time()  # capture immediately
        odom = self.get_odom()
        map_p = self.get_map_pose()

        with self._lock:
            self.seq_burst_time[seq] = ros_time
            self.seq_label[seq] = label
            self.seq_n_frames[seq] = n_frames
            self.current_label = label
            acc = self._burst_acc[seq]
            acc['odom'] = odom
            acc['map'] = map_p
            acc['n_frames'] = n_frames

        print(f"[BURST_START] seq={seq} label={label} n_frames={n_frames}")

    def _handle_ftm(self, parts):
        # FTM_F,<seq>,<frame_idx>,<rtt_ps>,<t1_ps>,<t2_ps>,<t3_ps>,<t4_ps>,<rssi_dbm>,<label>
        if len(parts) < 9:
            return
        try:
            seq = int(parts[1])
            frame_idx = int(parts[2])
            rtt_ps = int(parts[3])
            t1_ps = int(parts[4])
            t2_ps = int(parts[5])
            t3_ps = int(parts[6])
            t4_ps = int(parts[7])
            rssi_dbm = float(parts[8])
            label = parts[9] if len(parts) > 9 else self.seq_label.get(seq, 'UNKNOWN')
        except (ValueError, IndexError):
            return

        distance_m = rtt_ps * SPEED_OF_LIGHT / (2.0 * 1e12)

        with self._lock:
            btime = self.seq_burst_time.get(seq, (0, 0))
        odom = self.get_odom()
        map_p = self.get_map_pose()

        row = [
            seq, frame_idx, label,
            btime[0], btime[1],
            rtt_ps, round(distance_m, 6),
            t1_ps, t2_ps, t3_ps, t4_ps,
            rssi_dbm,
            odom[0], odom[1], odom[2], odom[3], odom[4],
            map_p[0], map_p[1], map_p[2],
        ]
        self._ftm_w.writerow(row)
        self._ftm_f.flush()

        with self._lock:
            acc = self._burst_acc[seq]
            acc['ftm_count'] += 1
            acc['rtt_list'].append(rtt_ps)
            acc['rssi_list'].append(rssi_dbm)

    def _handle_csi(self, parts):
        # CSI,<seq>,<rssi_dbm>,<noise_floor_dbm>,<n_sub>,<amp0>,...,<ampN-1>,<label>
        if len(parts) < 6:
            return
        try:
            seq = int(parts[1])
            rssi_dbm = float(parts[2])
            noise_floor_dbm = float(parts[3])
            n_sub = int(parts[4])
        except (ValueError, IndexError):
            return

        # Amplitudes are parts[5] through parts[5+n_sub-1]
        # Label (if present) is parts[5+n_sub]
        amp_parts = parts[5:5 + n_sub]
        label_idx = 5 + n_sub
        label = parts[label_idx] if len(parts) > label_idx else self.seq_label.get(seq, 'UNKNOWN')

        amps = []
        for a in amp_parts:
            try:
                amps.append(float(a))
            except ValueError:
                amps.append(0.0)
        amps = (amps + [0.0] * 52)[:52]  # pad/trim to exactly 52

        with self._lock:
            btime = self.seq_burst_time.get(seq, (0, 0))
        odom = self.get_odom()
        map_p = self.get_map_pose()

        row = (
            [seq, label, btime[0], btime[1], rssi_dbm, noise_floor_dbm, n_sub]
            + amps
            + [odom[0], odom[1], odom[2], odom[3], odom[4],
               map_p[0], map_p[1], map_p[2]]
        )
        self._csi_w.writerow(row)
        self._csi_f.flush()

        with self._lock:
            self._burst_acc[seq]['csi_count'] += 1

    def _handle_label(self, parts):
        if len(parts) >= 2:
            with self._lock:
                self.current_label = parts[1]
            print(f"[LABEL] -> {parts[1]}")

    def finalize(self):
        """Write burst summary and close files."""
        with self._lock:
            burst_items = sorted(self._burst_acc.items())

        for seq, acc in burst_items:
            label = self.seq_label.get(seq, 'UNKNOWN')
            btime = self.seq_burst_time.get(seq, (0, 0))
            rtt_list = acc['rtt_list']
            rssi_list = acc['rssi_list']
            mean_rtt = sum(rtt_list) / len(rtt_list) if rtt_list else 0.0
            mean_dist = mean_rtt * SPEED_OF_LIGHT / (2.0 * 1e12) if mean_rtt else 0.0
            mean_rssi = sum(rssi_list) / len(rssi_list) if rssi_list else 0.0
            o = acc['odom']   # (sec, ns, x, y, yaw)
            m = acc['map']    # (x, y, yaw)

            row = [
                seq, label,
                btime[0], btime[1],
                acc['n_frames'], acc['ftm_count'], acc['csi_count'],
                round(mean_rtt, 2), round(mean_dist, 6), round(mean_rssi, 2),
                o[2], o[3], o[4],
                m[0], m[1], m[2],
            ]
            self._burst_w.writerow(row)

        self._burst_f.flush()
        self._ftm_f.close()
        self._csi_f.close()
        self._burst_f.close()

    def save_excel(self, xlsx_path):
        if not OPENPYXL_AVAILABLE:
            print("\nOpenpyxl not available — Excel file NOT created.")
            print("Install with:  pip install openpyxl")
            print(f"CSVs are in:   {self.run_dir}")
            return

        wb = openpyxl.Workbook()

        def csv_to_sheet(csv_path, sheet):
            if not os.path.exists(csv_path):
                return
            with open(csv_path, newline='') as f:
                for row in csv.reader(f):
                    sheet.append(row)

        ws_meta = wb.active
        ws_meta.title = 'metadata'
        meta_path = os.path.join(self.run_dir, 'metadata.json')
        if os.path.exists(meta_path):
            with open(meta_path) as f:
                meta = json.load(f)
            ws_meta.append(['key', 'value'])
            for k, v in meta.items():
                ws_meta.append([str(k), str(v)])

        csv_to_sheet(os.path.join(self.run_dir, 'burst_summary.csv'),
                     wb.create_sheet('burst_summary'))
        csv_to_sheet(os.path.join(self.run_dir, 'ftm_frames.csv'),
                     wb.create_sheet('ftm_frames'))
        csv_to_sheet(os.path.join(self.run_dir, 'csi_packets.csv'),
                     wb.create_sheet('csi_packets'))

        wb.save(xlsx_path)
        print(f"\nExcel saved: {xlsx_path}")


# =============================================================================
# ROS2 Node wrapper
# =============================================================================

if ROS_AVAILABLE:
    class RttCsiOdomLoggerNode(Node):
        def __init__(self):
            super().__init__('rtt_csi_odom_logger')

            self.declare_parameter('serial_port', '/dev/ttyACM0')
            self.declare_parameter('baud_rate', 115200)
            self.declare_parameter('output_dir', os.path.expanduser('~/phoenix_data'))
            self.declare_parameter('odom_topic', '/diff_cont/odom')
            self.declare_parameter('fixed_frame', 'map')
            self.declare_parameter('base_frame', 'base_link')
            self.declare_parameter('notes', '')
            self.declare_parameter('map_file', '/home/phoenix/ros2_ws/maps/my_map.yaml')

            serial_port = self.get_parameter('serial_port').value
            baud_rate = int(self.get_parameter('baud_rate').value)
            output_dir = self.get_parameter('output_dir').value
            odom_topic = self.get_parameter('odom_topic').value
            self.fixed_frame = self.get_parameter('fixed_frame').value
            self.base_frame = self.get_parameter('base_frame').value
            notes = self.get_parameter('notes').value
            map_file = self.get_parameter('map_file').value

            # State
            self._odom_msg = None
            self._map_pose = None   # (x, y, yaw)
            self._odom_lock = threading.Lock()
            self._map_lock = threading.Lock()

            # TF
            self.tf_buffer = tf2_ros.Buffer()
            self.tf_listener = tf2_ros.TransformListener(self.tf_buffer, self)

            # Odom subscriber
            self.create_subscription(Odometry, odom_topic, self._odom_cb, 10)

            # Timer to refresh TF at 10 Hz
            self.create_timer(0.1, self._update_map_pose)

            # Build run directory
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.run_name = f'run_{ts}'
            run_dir = os.path.join(os.path.expanduser(output_dir), self.run_name)

            # Create logger
            self.dlogger = DataLogger(
                run_dir,
                get_ros_time_fn=self._ros_time,
                get_odom_fn=self._get_odom,
                get_map_pose_fn=self._get_map_pose,
            )

            # Save metadata
            meta = {
                "robot": "PHOENIX",
                "ros_distro": "jazzy",
                "ranging_initiator": "XIAO ESP32-S3 onboard PHOENIX",
                "target_ap": "stationary XIAO ESP32-S3 FTM_AP",
                "wifi_ssid": "FTM_AP",
                "wifi_channel": 6,
                "serial_port": serial_port,
                "serial_baud": baud_rate,
                "odom_topic": odom_topic,
                "fixed_frame": self.fixed_frame,
                "base_frame": self.base_frame,
                "notes": notes,
                "map_file": map_file,
                "run_name": self.run_name,
                "start_time_utc": datetime.now(timezone.utc).isoformat(),
            }
            with open(os.path.join(run_dir, 'metadata.json'), 'w') as f:
                json.dump(meta, f, indent=2)

            # Open serial
            try:
                self._serial = serial.Serial(serial_port, baud_rate, timeout=1.0)
                self.get_logger().info(f"Serial open: {serial_port} @ {baud_rate}")
            except serial.SerialException as e:
                self.get_logger().fatal(f"Cannot open serial {serial_port}: {e}")
                raise

            # Serial reader thread
            self._stop = False
            self._thread = threading.Thread(target=self._serial_reader, daemon=True)
            self._thread.start()

            self.xlsx_path = os.path.join(
                os.path.expanduser(output_dir),
                f'{self.run_name}.xlsx'
            )

            self.get_logger().info(f"Logger started. Output: {run_dir}")
            self.get_logger().info("Press Ctrl+C to stop and save Excel.")

        def _ros_time(self):
            t = self.get_clock().now()
            ns = t.nanoseconds
            return ns // 10**9, ns % 10**9

        def _odom_cb(self, msg: Odometry):
            with self._odom_lock:
                self._odom_msg = msg

        def _update_map_pose(self):
            try:
                tf = self.tf_buffer.lookup_transform(
                    self.fixed_frame, self.base_frame,
                    rclpy.time.Time(),
                    timeout=Duration(seconds=0.05),
                )
                t = tf.transform.translation
                q = tf.transform.rotation
                yaw = math.atan2(
                    2.0 * (q.w * q.z + q.x * q.y),
                    1.0 - 2.0 * (q.y * q.y + q.z * q.z),
                )
                with self._map_lock:
                    self._map_pose = (t.x, t.y, yaw)
            except Exception:
                pass  # TF not yet available — keep None

        def _get_odom(self):
            with self._odom_lock:
                msg = self._odom_msg
            if msg is None:
                return None, None, None, None, None
            q = msg.pose.pose.orientation
            yaw = math.atan2(
                2.0 * (q.w * q.z + q.x * q.y),
                1.0 - 2.0 * (q.y * q.y + q.z * q.z),
            )
            return (
                msg.header.stamp.sec,
                msg.header.stamp.nanosec,
                msg.pose.pose.position.x,
                msg.pose.pose.position.y,
                yaw,
            )

        def _get_map_pose(self):
            with self._map_lock:
                p = self._map_pose
            return p if p is not None else (None, None, None)

        def _serial_reader(self):
            while not self._stop:
                try:
                    raw = self._serial.readline()
                    if not raw:
                        continue
                    line = raw.decode('utf-8', errors='replace').strip()
                    if line:
                        self.dlogger.parse_line(line)
                except serial.SerialException:
                    break
                except Exception as e:
                    self.get_logger().warn(f"Parse error: {e}")

        def shutdown(self):
            self._stop = True
            self.dlogger.finalize()
            if self._serial.is_open:
                self._serial.close()
            self.dlogger.save_excel(self.xlsx_path)
            self.get_logger().info(f"Data saved to: {self.dlogger.run_dir}")


# =============================================================================
# Standalone mode (no ROS)
# =============================================================================

class StandaloneLogger:
    """Runs without ROS — no odom/TF sync, only serial data."""

    def __init__(self, serial_port, baud_rate, output_dir):
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.run_name = f'run_{ts}'
        run_dir = os.path.join(os.path.expanduser(output_dir), self.run_name)

        def no_ros_time():
            now = datetime.now(timezone.utc)
            epoch_ns = int(now.timestamp() * 1e9)
            return epoch_ns // 10**9, epoch_ns % 10**9

        def no_odom():
            return None, None, None, None, None

        def no_map():
            return None, None, None

        self.dlogger = DataLogger(run_dir, no_ros_time, no_odom, no_map)

        meta = {
            "robot": "PHOENIX",
            "mode": "standalone_no_ros",
            "serial_port": serial_port,
            "serial_baud": baud_rate,
            "run_name": self.run_name,
            "start_time_utc": datetime.now(timezone.utc).isoformat(),
        }
        with open(os.path.join(run_dir, 'metadata.json'), 'w') as f:
            json.dump(meta, f, indent=2)

        self._serial = serial.Serial(serial_port, baud_rate, timeout=1.0)
        self.xlsx_path = os.path.join(
            os.path.expanduser(output_dir), f'{self.run_name}.xlsx'
        )
        print(f"[Standalone] Logging to: {run_dir}")
        print("Press Ctrl+C to stop.")

    def run(self):
        try:
            while True:
                raw = self._serial.readline()
                if not raw:
                    continue
                line = raw.decode('utf-8', errors='replace').strip()
                if line:
                    self.dlogger.parse_line(line)
        except KeyboardInterrupt:
            pass
        finally:
            self.dlogger.finalize()
            self._serial.close()
            self.dlogger.save_excel(self.xlsx_path)
            print(f"Done. Data in: {self.dlogger.run_dir}")


# =============================================================================
# Entry points
# =============================================================================

def main_ros(args=None):
    rclpy.init(args=args)
    node = RttCsiOdomLoggerNode()
    try:
        rclpy.spin(node)
    except KeyboardInterrupt:
        pass
    finally:
        node.shutdown()
        node.destroy_node()
        rclpy.shutdown()


def main_standalone():
    parser = argparse.ArgumentParser(
        description='RTT/CSI logger (standalone, no ROS)'
    )
    parser.add_argument('--serial-port', default='/dev/ttyACM0')
    parser.add_argument('--baud', type=int, default=115200)
    parser.add_argument('--output-dir', default=os.path.expanduser('~/phoenix_data'))
    args = parser.parse_args()

    logger = StandaloneLogger(args.serial_port, args.baud, args.output_dir)
    logger.run()


# When invoked via 'ros2 run' or directly as a script
if __name__ == '__main__':
    if ROS_AVAILABLE and '--ros-args' not in sys.argv and len(sys.argv) > 1:
        # Has command-line args but not ROS-style → standalone mode
        main_standalone()
    elif ROS_AVAILABLE:
        main_ros()
    else:
        main_standalone()


# ROS2 entry point (used by 'ros2 run articubot_one rtt_csi_odom_logger')
def ros_main(args=None):
    main_ros(args)
